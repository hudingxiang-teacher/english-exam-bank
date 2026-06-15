# -*- coding: utf-8 -*-
"""中考英语真题题库 - 本地版 (完整版) 单文件 Flask 后端"""
import os, json, sqlite3, hashlib, secrets, csv, io, re, random
from datetime import datetime, timedelta
from functools import wraps
from contextlib import contextmanager
from flask import Flask, request, jsonify, send_from_directory, send_file

BASE = os.path.dirname(os.path.abspath(__file__))
# Render 持久化磁盘挂载点（免费版无持久化，数据重启后丢失）
RENDER_DATA_DIR = os.environ.get("RENDER_DATA_DIR", "")
if RENDER_DATA_DIR and os.path.isdir(RENDER_DATA_DIR):
    DATA_DIR = RENDER_DATA_DIR
else:
    DATA_DIR = os.path.join(BASE, "data")
DB_PATH = os.path.join(DATA_DIR, "exam.db")
UPLOAD_DIR = os.path.join(DATA_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)

# 云部署：环境变量配置
DATABASE_URL = os.environ.get("DATABASE_URL", "").strip()
SECRET_KEY = os.environ.get("SECRET_KEY", "exam-bank-default-secret-change-me")
PORT = int(os.environ.get("PORT", 5000))
USE_POSTGRES = bool(DATABASE_URL) and DATABASE_URL.startswith("postgres")

# PostgreSQL 适配
if USE_POSTGRES:
    try:
        import psycopg2
        import psycopg2.extras
    except ImportError:
        print("[!] 云端部署需要 psycopg2-binary：pip install psycopg2-binary")
        USE_POSTGRES = False
    # Render 提供 postgres://，但 psycopg2 需要 postgresql://
    if DATABASE_URL.startswith("postgres://"):
        DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

app = Flask(__name__, static_folder="static", static_url_path="")
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024
app.config["SECRET_KEY"] = SECRET_KEY

# ---------- DB 适配层（SQLite 本地 / PostgreSQL 云端） ----------
class _SqliteRow(dict):
    """让 sqlite3.Row 支持 dict 风格访问（与 psycopg2 行为一致）"""
    def __getitem__(self, k):
        v = super().__getitem__(k)
        return v

class _CursorWrap:
    """统一 cursor 接口：fetchone/fetchall 返回 dict（与 sqlite3.Row 兼容）"""
    def __init__(self, cursor, is_pg=False, returning_id=False):
        self._c = cursor
        self._pg = is_pg
        self._returning_id = returning_id
    def fetchone(self):
        if self._pg:
            row = self._c.fetchone()
            return dict(row) if row else None
        row = self._c.fetchone()
        return dict(row) if row else None
    def fetchall(self):
        if self._pg:
            return [dict(r) for r in self._c.fetchall()]
        return [dict(r) for r in self._c.fetchall()]
    @property
    def lastrowid(self):
        if self._pg and self._returning_id:
            row = self._c.fetchone()
            return row["id"] if row else None
        return self._c.lastrowid
    def __iter__(self):
        if self._pg:
            for r in self._c:
                yield dict(r)
        else:
            for r in self._c:
                yield dict(r)


class _ConnWrap:
    """统一 Connection 接口：execute 自动 _q 转换；fetchall 返回 dict"""
    def __init__(self, conn):
        self._c = conn
        self._pg = USE_POSTGRES
    def execute(self, sql, args=()):
        sql = _q(sql)
        # PostgreSQL: INSERT 自动加 RETURNING id，让 lastrowid 可用
        returning_id = False
        if self._pg and sql.lstrip().upper().startswith("INSERT") and "RETURNING" not in sql.upper():
            # 检查是否有自增 id 列（heuristic：表名 + 返回 id）
            sql = sql.rstrip().rstrip(";") + " RETURNING id"
            returning_id = True
        if self._pg:
            cur = self._c.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur.execute(sql, args)
            return _CursorWrap(cur, is_pg=True, returning_id=returning_id)
        else:
            return self._c.execute(sql, args)
    def executescript(self, sql_script):
        """执行多语句脚本（CREATE TABLE 等）。"""
        if self._pg:
            with self._c.cursor() as cur:
                cur.execute(sql_script)
        else:
            self._c.executescript(sql_script)
    def commit(self): self._c.commit()
    def rollback(self): self._c.rollback()
    def close(self): self._c.close()
    def cursor(self):
        return self._c.cursor()


_DB_INITIALIZED = False

@contextmanager
def db():
    """数据库上下文管理器。自动选择 SQLite / PostgreSQL。"""
    global _DB_INITIALIZED
    if not _DB_INITIALIZED:
        # 第一次调用前自动 init_db（gunicorn worker 启动时也会跑）
        _init_db_internal()
        _DB_INITIALIZED = True
    if USE_POSTGRES:
        conn = psycopg2.connect(DATABASE_URL)
    else:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
    wrapped = _ConnWrap(conn)
    try:
        yield wrapped
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def _init_db_internal():
    """实际执行 schema 初始化"""
    if USE_POSTGRES:
        conn = psycopg2.connect(DATABASE_URL)
    else:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
    try:
        # 调 init_db 但跳过它内部的 db() 调用（避免循环）
        if USE_POSTGRES:
            schema_script = """
            CREATE TABLE IF NOT EXISTS classes (
                id SERIAL PRIMARY KEY, name TEXT NOT NULL, grade TEXT,
                teacher_id INTEGER, created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(name, grade)
            );
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY, username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL, name TEXT,
                role TEXT DEFAULT 'student', class_id INTEGER, student_no TEXT,
                active INTEGER DEFAULT 1, created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS questions (
                id SERIAL PRIMARY KEY, source TEXT, year INTEGER, region TEXT,
                qtype TEXT NOT NULL, difficulty INTEGER DEFAULT 3,
                stem TEXT NOT NULL, options TEXT, answer TEXT NOT NULL,
                analysis TEXT, tags TEXT, kp TEXT, score INTEGER DEFAULT 2,
                active INTEGER DEFAULT 1, created_by INTEGER,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS knowledge_points (
                id SERIAL PRIMARY KEY, code TEXT UNIQUE NOT NULL, name TEXT NOT NULL,
                category TEXT, description TEXT, created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS papers (
                id SERIAL PRIMARY KEY, name TEXT NOT NULL, description TEXT,
                duration_minutes INTEGER DEFAULT 120, total_score INTEGER DEFAULT 100,
                qtype_filter TEXT, question_count INTEGER DEFAULT 0,
                published INTEGER DEFAULT 1, class_id INTEGER, created_by INTEGER,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS paper_questions (
                id SERIAL PRIMARY KEY, paper_id INTEGER NOT NULL, question_id INTEGER NOT NULL,
                order_index INTEGER DEFAULT 0, score INTEGER DEFAULT 2, UNIQUE(paper_id, question_id)
            );
            CREATE TABLE IF NOT EXISTS attempts (
                id SERIAL PRIMARY KEY, user_id INTEGER NOT NULL, mode TEXT NOT NULL,
                paper_id INTEGER, status TEXT DEFAULT 'in_progress',
                total INTEGER DEFAULT 0, correct INTEGER DEFAULT 0, qcount INTEGER DEFAULT 0,
                score_earned REAL DEFAULT 0, max_score REAL DEFAULT 0,
                started_at TEXT DEFAULT CURRENT_TIMESTAMP, submitted_at TEXT,
                duration_seconds INTEGER, qids TEXT
            );
            CREATE TABLE IF NOT EXISTS answers (
                id SERIAL PRIMARY KEY, attempt_id INTEGER NOT NULL, user_id INTEGER NOT NULL,
                question_id INTEGER NOT NULL, user_answer TEXT,
                is_correct INTEGER DEFAULT 0, score_earned REAL DEFAULT 0,
                time_spent INTEGER DEFAULT 0, answered_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS wrong_book (
                id SERIAL PRIMARY KEY, user_id INTEGER NOT NULL, question_id INTEGER NOT NULL,
                wrong_count INTEGER DEFAULT 1, last_wrong_at TEXT DEFAULT CURRENT_TIMESTAMP,
                mastered INTEGER DEFAULT 0, UNIQUE(user_id, question_id)
            );
            CREATE TABLE IF NOT EXISTS sessions (
                token TEXT PRIMARY KEY, user_id INTEGER NOT NULL,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            """
        else:
            schema_script = """
            CREATE TABLE IF NOT EXISTS classes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL, grade TEXT, teacher_id INTEGER,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP, UNIQUE(name, grade)
            );
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL, password TEXT NOT NULL, name TEXT,
                role TEXT DEFAULT "student", class_id INTEGER, student_no TEXT,
                active INTEGER DEFAULT 1, created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS questions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source TEXT, year INTEGER, region TEXT,
                qtype TEXT NOT NULL, difficulty INTEGER DEFAULT 3,
                stem TEXT NOT NULL, options TEXT, answer TEXT NOT NULL,
                analysis TEXT, tags TEXT, kp TEXT, score INTEGER DEFAULT 2,
                active INTEGER DEFAULT 1, created_by INTEGER,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS knowledge_points (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT UNIQUE NOT NULL, name TEXT NOT NULL, category TEXT, description TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS papers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL, description TEXT,
                duration_minutes INTEGER DEFAULT 120, total_score INTEGER DEFAULT 100,
                qtype_filter TEXT, question_count INTEGER DEFAULT 0,
                published INTEGER DEFAULT 1, class_id INTEGER, created_by INTEGER,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS paper_questions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                paper_id INTEGER NOT NULL, question_id INTEGER NOT NULL,
                order_index INTEGER DEFAULT 0, score INTEGER DEFAULT 2, UNIQUE(paper_id, question_id)
            );
            CREATE TABLE IF NOT EXISTS attempts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL, mode TEXT NOT NULL, paper_id INTEGER,
                status TEXT DEFAULT "in_progress",
                total INTEGER DEFAULT 0, correct INTEGER DEFAULT 0, qcount INTEGER DEFAULT 0,
                score_earned REAL DEFAULT 0, max_score REAL DEFAULT 0,
                started_at TEXT DEFAULT CURRENT_TIMESTAMP, submitted_at TEXT,
                duration_seconds INTEGER, qids TEXT
            );
            CREATE TABLE IF NOT EXISTS answers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                attempt_id INTEGER NOT NULL, user_id INTEGER NOT NULL, question_id INTEGER NOT NULL,
                user_answer TEXT, is_correct INTEGER DEFAULT 0,
                score_earned REAL DEFAULT 0, time_spent INTEGER DEFAULT 0,
                answered_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS wrong_book (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL, question_id INTEGER NOT NULL,
                wrong_count INTEGER DEFAULT 1, last_wrong_at TEXT DEFAULT CURRENT_TIMESTAMP,
                mastered INTEGER DEFAULT 0, UNIQUE(user_id, question_id)
            );
            CREATE TABLE IF NOT EXISTS sessions (
                token TEXT PRIMARY KEY, user_id INTEGER NOT NULL,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            """
        wrapped_init = _ConnWrap(conn)
        if USE_POSTGRES:
            with conn.cursor() as cur:
                cur.execute(schema_script)
        else:
            conn.executescript(schema_script)
        # migration
        for col, default in [('passage_id', 0), ('sub_idx', 0), ('unit_no', 0)]:
            try:
                wrapped_init.execute(f"ALTER TABLE questions ADD COLUMN {col} INTEGER DEFAULT {default}")
            except Exception:
                pass
        try:
            wrapped_init.execute("CREATE INDEX IF NOT EXISTS idx_q_unit ON questions(unit_no, sub_idx)")
        except Exception:
            pass
        # 默认账号（用包装层）
        # RealDictCursor 由 _ConnWrap 内部使用
        wrapped = _ConnWrap(conn)
        if not wrapped.execute("SELECT id FROM users WHERE username=?", ("admin",)).fetchone():
            wrapped.execute("INSERT INTO users(username,password,name,role) VALUES(?,?,?,?)",
                ("admin", hash_pw("admin123"), "系统管理员", "admin"))
        if not wrapped.execute("SELECT id FROM users WHERE username=?", ("teacher",)).fetchone():
            wrapped.execute("INSERT INTO users(username,password,name,role) VALUES(?,?,?,?)",
                ("teacher", hash_pw("teacher123"), "示例教师", "teacher"))
        if not wrapped.execute("SELECT id FROM users WHERE username=?", ("student",)).fetchone():
            wrapped.execute("INSERT INTO users(username,password,name,role,class_id,student_no) VALUES(?,?,?,?,?,?)",
                ("student", hash_pw("123456"), "示例学生", "student", 1, "S001"))
        if not wrapped.execute("SELECT id FROM users WHERE username=?", ("student02",)).fetchone():
            wrapped.execute("INSERT INTO users(username,password,name,role,class_id,student_no) VALUES(?,?,?,?,?,?)",
                ("student02", hash_pw("123456"), "示例学生2", "student", 1, "S002"))
        if not wrapped.execute("SELECT id FROM classes").fetchone():
            wrapped.execute("INSERT INTO classes(name,grade) VALUES(?,?)", ("三年一班", "初三"))
            wrapped.execute("INSERT INTO classes(name,grade) VALUES(?,?)", ("三年二班", "初三"))
        default_kps = [
            ("KP_GRAMMAR_TENSE", "时态", "语法"),
            ("KP_GRAMMAR_ARTICLE", "冠词", "语法"),
            ("KP_GRAMMAR_SUBJUNCTIVE", "虚拟语气/条件句", "语法"),
            ("KP_GRAMMAR_CLAUSE_RELATIVE", "定语从句", "语法"),
            ("KP_GRAMMAR_THERE_BE", "There be 句型", "语法"),
            ("KP_GRAMMAR_QUANTIFIER", "数量词", "语法"),
            ("KP_VOCAB_NOUN", "名词", "词汇"),
            ("KP_VOCAB_VERB", "动词", "词汇"),
            ("KP_VOCAB_PHRASE", "短语搭配", "词汇"),
            ("KP_READING_DETAIL", "阅读-细节理解", "阅读"),
            ("KP_READING_MAIN_IDEA", "阅读-主旨大意", "阅读"),
            ("KP_READING_INFERENCE", "阅读-推理判断", "阅读"),
            ("KP_CLOZE", "完形填空", "综合"),
            ("KP_TRANSLATION", "翻译", "写作"),
            ("KP_WRITING", "写作", "写作"),
        ]
        for code, name, cat in default_kps:
            if not wrapped.execute("SELECT id FROM knowledge_points WHERE code=?", (code,)).fetchone():
                wrapped.execute("INSERT INTO knowledge_points(code,name,category) VALUES(?,?,?)", (code, name, cat))
        conn.commit()
    finally:
        conn.close()


def _q(sql):
    """SQL 占位符适配：sqlite3 用 ?, PostgreSQL 用 %s"""
    return sql.replace("?", "%s") if USE_POSTGRES else sql

def hash_pw(p):
    return hashlib.sha256((p + "salt_exam_2024").encode()).hexdigest()

def init_db():
    with db() as c:
        # PostgreSQL 用 SERIAL；SQLite 用 INTEGER PRIMARY KEY AUTOINCREMENT
        if USE_POSTGRES:
            schema = """
            CREATE TABLE IF NOT EXISTS classes (
                id SERIAL PRIMARY KEY, name TEXT NOT NULL, grade TEXT,
                teacher_id INTEGER, created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(name, grade)
            );
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY, username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL, name TEXT,
                role TEXT DEFAULT 'student', class_id INTEGER, student_no TEXT,
                active INTEGER DEFAULT 1, created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS questions (
                id SERIAL PRIMARY KEY,
                source TEXT, year INTEGER, region TEXT,
                qtype TEXT NOT NULL, difficulty INTEGER DEFAULT 3,
                stem TEXT NOT NULL, options TEXT, answer TEXT NOT NULL,
                analysis TEXT, tags TEXT, kp TEXT, score INTEGER DEFAULT 2,
                active INTEGER DEFAULT 1, created_by INTEGER,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS knowledge_points (
                id SERIAL PRIMARY KEY, code TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL, category TEXT, description TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS papers (
                id SERIAL PRIMARY KEY, name TEXT NOT NULL, description TEXT,
                duration_minutes INTEGER DEFAULT 120, total_score INTEGER DEFAULT 100,
                qtype_filter TEXT, question_count INTEGER DEFAULT 0,
                published INTEGER DEFAULT 1, class_id INTEGER, created_by INTEGER,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS paper_questions (
                id SERIAL PRIMARY KEY, paper_id INTEGER NOT NULL, question_id INTEGER NOT NULL,
                order_index INTEGER DEFAULT 0, score INTEGER DEFAULT 2,
                UNIQUE(paper_id, question_id)
            );
            CREATE TABLE IF NOT EXISTS attempts (
                id SERIAL PRIMARY KEY, user_id INTEGER NOT NULL, mode TEXT NOT NULL,
                paper_id INTEGER, status TEXT DEFAULT 'in_progress',
                total INTEGER DEFAULT 0, correct INTEGER DEFAULT 0, qcount INTEGER DEFAULT 0,
                score_earned REAL DEFAULT 0, max_score REAL DEFAULT 0,
                started_at TEXT DEFAULT CURRENT_TIMESTAMP, submitted_at TEXT,
                duration_seconds INTEGER, qids TEXT
            );
            CREATE TABLE IF NOT EXISTS answers (
                id SERIAL PRIMARY KEY, attempt_id INTEGER NOT NULL, user_id INTEGER NOT NULL,
                question_id INTEGER NOT NULL, user_answer TEXT,
                is_correct INTEGER DEFAULT 0, score_earned REAL DEFAULT 0,
                time_spent INTEGER DEFAULT 0, answered_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS wrong_book (
                id SERIAL PRIMARY KEY, user_id INTEGER NOT NULL, question_id INTEGER NOT NULL,
                wrong_count INTEGER DEFAULT 1, last_wrong_at TEXT DEFAULT CURRENT_TIMESTAMP,
                mastered INTEGER DEFAULT 0, UNIQUE(user_id, question_id)
            );
            CREATE TABLE IF NOT EXISTS sessions (
                token TEXT PRIMARY KEY, user_id INTEGER NOT NULL,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            """
        else:
            schema = """
        CREATE TABLE IF NOT EXISTS classes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL, grade TEXT, teacher_id INTEGER,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP, UNIQUE(name, grade)
        );
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL, password TEXT NOT NULL, name TEXT,
            role TEXT DEFAULT "student", class_id INTEGER, student_no TEXT,
            active INTEGER DEFAULT 1, created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS questions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source TEXT, year INTEGER, region TEXT,
            qtype TEXT NOT NULL, difficulty INTEGER DEFAULT 3,
            stem TEXT NOT NULL, options TEXT, answer TEXT NOT NULL,
            analysis TEXT, tags TEXT, kp TEXT, score INTEGER DEFAULT 2,
            active INTEGER DEFAULT 1, created_by INTEGER,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS knowledge_points (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE NOT NULL, name TEXT NOT NULL, category TEXT, description TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS papers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL, description TEXT,
            duration_minutes INTEGER DEFAULT 120, total_score INTEGER DEFAULT 100,
            qtype_filter TEXT, question_count INTEGER DEFAULT 0,
            published INTEGER DEFAULT 1, class_id INTEGER, created_by INTEGER,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS paper_questions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            paper_id INTEGER NOT NULL, question_id INTEGER NOT NULL,
            order_index INTEGER DEFAULT 0, score INTEGER DEFAULT 2,
            UNIQUE(paper_id, question_id)
        );
        CREATE TABLE IF NOT EXISTS attempts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL, mode TEXT NOT NULL, paper_id INTEGER,
            status TEXT DEFAULT "in_progress",
            total INTEGER DEFAULT 0, correct INTEGER DEFAULT 0, qcount INTEGER DEFAULT 0,
            score_earned REAL DEFAULT 0, max_score REAL DEFAULT 0,
            started_at TEXT DEFAULT CURRENT_TIMESTAMP, submitted_at TEXT,
            duration_seconds INTEGER, qids TEXT
        );
        CREATE TABLE IF NOT EXISTS answers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            attempt_id INTEGER NOT NULL, user_id INTEGER NOT NULL, question_id INTEGER NOT NULL,
            user_answer TEXT, is_correct INTEGER DEFAULT 0,
            score_earned REAL DEFAULT 0, time_spent INTEGER DEFAULT 0,
            answered_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS wrong_book (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL, question_id INTEGER NOT NULL,
            wrong_count INTEGER DEFAULT 1, last_wrong_at TEXT DEFAULT CURRENT_TIMESTAMP,
            mastered INTEGER DEFAULT 0, UNIQUE(user_id, question_id)
        );
        CREATE TABLE IF NOT EXISTS sessions (
            token TEXT PRIMARY KEY,
            user_id INTEGER NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        """
        c.executescript(schema)
        # schema migration
        for col, default in [('passage_id', 0), ('sub_idx', 0), ('unit_no', 0)]:
            try:
                c.execute(f"ALTER TABLE questions ADD COLUMN {col} INTEGER DEFAULT {default}")
            except Exception:
                pass
        try:
            c.execute("CREATE INDEX IF NOT EXISTS idx_q_unit ON questions(unit_no, sub_idx)")
        except Exception:
            pass
        # 默认账号
        if not c.execute("SELECT id FROM users WHERE username=?", ("admin",)).fetchone():
            c.execute("INSERT INTO users(username,password,name,role) VALUES(?,?,?,?)",
                ("admin", hash_pw("admin123"), "系统管理员", "admin"))
        if not c.execute("SELECT id FROM users WHERE username=?", ("teacher",)).fetchone():
            c.execute("INSERT INTO users(username,password,name,role) VALUES(?,?,?,?)",
                ("teacher", hash_pw("teacher123"), "示例教师", "teacher"))
        if not c.execute("SELECT id FROM users WHERE username=?", ("student",)).fetchone():
            c.execute("INSERT INTO users(username,password,name,role,class_id,student_no) VALUES(?,?,?,?,?,?)",
                ("student", hash_pw("123456"), "示例学生", "student", 1, "S001"))
        if not c.execute("SELECT id FROM users WHERE username=?", ("student02",)).fetchone():
            c.execute("INSERT INTO users(username,password,name,role,class_id,student_no) VALUES(?,?,?,?,?,?)",
                ("student02", hash_pw("123456"), "示例学生2", "student", 1, "S002"))
        if not c.execute("SELECT id FROM classes").fetchone():
            c.execute("INSERT INTO classes(name,grade) VALUES(?,?)", ("三年一班", "初三"))
            c.execute("INSERT INTO classes(name,grade) VALUES(?,?)", ("三年二班", "初三"))
        default_kps = [
                ("KP_GRAMMAR_TENSE", "时态", "语法"),
                ("KP_GRAMMAR_ARTICLE", "冠词", "语法"),
                ("KP_GRAMMAR_SUBJUNCTIVE", "虚拟语气/条件句", "语法"),
                ("KP_GRAMMAR_CLAUSE_RELATIVE", "定语从句", "语法"),
                ("KP_GRAMMAR_THERE_BE", "There be 句型", "语法"),
                ("KP_GRAMMAR_QUANTIFIER", "数量词", "语法"),
                ("KP_VOCAB_NOUN", "名词", "词汇"),
                ("KP_VOCAB_VERB", "动词", "词汇"),
                ("KP_VOCAB_PHRASE", "短语搭配", "词汇"),
                ("KP_READING_DETAIL", "阅读-细节理解", "阅读"),
                ("KP_READING_MAIN_IDEA", "阅读-主旨大意", "阅读"),
                ("KP_READING_INFERENCE", "阅读-推理判断", "阅读"),
                ("KP_CLOZE", "完形填空", "综合"),
                ("KP_TRANSLATION", "翻译", "写作"),
                ("KP_WRITING", "写作", "写作"),
        ]
        for code, name, cat in default_kps:
            if not c.execute("SELECT id FROM knowledge_points WHERE code=?", (code,)).fetchone():
                c.execute("INSERT INTO knowledge_points(code,name,category) VALUES(?,?,?)", (code, name, cat))

def _is_duplicate_key_error(e):
    """判断是否为唯一约束冲突（兼容 SQLite 和 PostgreSQL）"""
    if USE_POSTGRES:
        try:
            import psycopg2.errors
            return isinstance(e, psycopg2.errors.UniqueViolation)
        except Exception:
            return "unique" in str(e).lower() or "duplicate" in str(e).lower()
    return isinstance(e, sqlite3.IntegrityError)

# ---------- AUTH ----------
def current_user():
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        token = auth[7:]
        with db() as c:
            row = c.execute("SELECT u.* FROM sessions s JOIN users u ON u.id=s.user_id WHERE s.token=?", (token,)).fetchone()
            return row
    return None

def login_required(f):
    @wraps(f)
    def w(*a, **k):
        u = current_user()
        if not u: return jsonify({"error": "请先登录"}), 401
        request.user = u
        return f(*a, **k)
    return w

def teacher_required(f):
    @wraps(f)
    def w(*a, **k):
        u = current_user()
        if not u: return jsonify({"error": "请先登录"}), 401
        if u["role"] not in ("admin", "teacher"): return jsonify({"error": "需要教师或管理员权限"}), 403
        request.user = u
        return f(*a, **k)
    return w

# ---------- ROUTES ----------
@app.route("/")
def index():
    return send_from_directory("static", "index.html")

@app.route("/api/login", methods=["POST"])
def login():
    d = request.get_json() or {}
    u = (d.get("username") or "").strip()
    p = d.get("password") or ""
    if not u or not p: return jsonify({"error": "用户名和密码必填"}), 400
    with db() as c:
        row = c.execute("SELECT * FROM users WHERE username=? AND active=1", (u,)).fetchone()
        if not row or row["password"] != hash_pw(p):
            return jsonify({"error": "用户名或密码错误"}), 401
        token = secrets.token_hex(32)
        c.execute("INSERT INTO sessions(token,user_id) VALUES(?,?)", (token, row["id"]))
        cls = None
        if row["class_id"]:
            cls = dict(c.execute("SELECT * FROM classes WHERE id=?", (row["class_id"],)).fetchone() or {})
        user = dict(row)
        user.pop("password", None)
        user["class_info"] = cls
        return jsonify({"token": token, "user": user})

@app.route("/api/logout", methods=["POST"])
@login_required
def logout():
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        with db() as c:
            c.execute("DELETE FROM sessions WHERE token=?", (auth[7:],))
    return jsonify({"ok": True})

@app.route("/api/me")
@login_required
def me():
    u = dict(request.user)
    u.pop("password", None)
    if u.get("class_id"):
        with db() as c:
            cls = c.execute("SELECT * FROM classes WHERE id=?", (u["class_id"],)).fetchone()
            u["class_info"] = dict(cls) if cls else None
    else:
        u["class_info"] = None
    return jsonify(u)

# ---------- 班级 ----------
@app.route("/api/classes", methods=["GET", "POST"])
@teacher_required
def classes():
    if request.method == "GET":
        with db() as c:
            rows = [dict(r) for r in c.execute("SELECT * FROM classes ORDER BY id").fetchall()]
            for r in rows:
                r["student_count"] = c.execute("SELECT COUNT(*) c FROM users WHERE class_id=? AND role='student' AND active=1", (r["id"],)).fetchone()["c"]
        return jsonify(rows)
    d = request.get_json() or {}
    with db() as c:
        try:
            cur = c.execute("INSERT INTO classes(name,grade) VALUES(?,?)", (d["name"], d.get("grade", "")))
            return jsonify({"ok": True, "id": cur.lastrowid})
        except Exception as e:
            if _is_duplicate_key_error(e):
                return jsonify({"error": "班级已存在"}), 400
            raise

@app.route("/api/classes/<int:cid>", methods=["PATCH", "DELETE"])
@teacher_required
def class_manage(cid):
    with db() as c:
        if request.method == "DELETE":
            n = c.execute("SELECT COUNT(*) c FROM users WHERE class_id=?", (cid,)).fetchone()["c"]
            if n > 0: return jsonify({"error": f"该班级有 {n} 名学生，无法删除"}), 400
            c.execute("DELETE FROM classes WHERE id=?", (cid,))
            return jsonify({"ok": True})
        d = request.get_json() or {}
        if "name" in d: c.execute("UPDATE classes SET name=? WHERE id=?", (d["name"], cid))
        if "grade" in d: c.execute("UPDATE classes SET grade=? WHERE id=?", (d["grade"], cid))
        return jsonify({"ok": True})

# ---------- 用户管理 ----------
@app.route("/api/users", methods=["GET"])
@teacher_required
def users_list():
    role = request.args.get("role")
    class_id = request.args.get("class_id")
    keyword = request.args.get("q", "").strip()
    with db() as c:
        sql = "SELECT u.id, u.username, u.name, u.role, u.class_id, u.student_no, u.active, u.created_at, c.name AS class_name, c.grade AS class_grade FROM users u LEFT JOIN classes c ON c.id=u.class_id WHERE 1=1"
        args = []
        if role: sql += " AND u.role=?"; args.append(role)
        if class_id: sql += " AND u.class_id=?"; args.append(class_id)
        if keyword: sql += " AND (u.username LIKE ? OR u.name LIKE ? OR u.student_no LIKE ?)"; args.extend([f"%{keyword}%"]*3)
        sql += " ORDER BY u.role, u.id"
        rows = [dict(r) for r in c.execute(sql, args).fetchall()]
        for r in rows: r.pop("password", None)
    return jsonify(rows)

@app.route("/api/users", methods=["POST"])
@teacher_required
def create_user():
    d = request.get_json() or {}
    if not d.get("username") or not d.get("password"):
        return jsonify({"error": "用户名和密码必填"}), 400
    if d.get("role") == "admin" and request.user["role"] != "admin":
        return jsonify({"error": "只有管理员能创建管理员"}), 403
    with db() as c:
        try:
            cur = c.execute("INSERT INTO users(username,password,name,role,class_id,student_no) VALUES(?,?,?,?,?,?)",
                (d["username"], hash_pw(d["password"]), d.get("name", ""), d.get("role", "student"), d.get("class_id") or None, d.get("student_no", "")))
            return jsonify({"ok": True, "id": cur.lastrowid})
        except Exception as e:
            if _is_duplicate_key_error(e):
                return jsonify({"error": "用户名已存在"}), 400
            raise

@app.route("/api/users/<int:uid>", methods=["GET", "PATCH", "DELETE"])
@teacher_required
def user_manage(uid):
    with db() as c:
        u = c.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
        if not u: return jsonify({"error": "用户不存在"}), 404
        if request.method == "GET":
            d = dict(u); d.pop("password", None)
            return jsonify(d)
        if request.method == "DELETE":
            if u["id"] == request.user["id"]: return jsonify({"error": "不能删除自己"}), 400
            if u["role"] == "admin" and request.user["role"] != "admin": return jsonify({"error": "无权删除管理员"}), 403
            c.execute("UPDATE users SET active=0 WHERE id=?", (uid,))
            return jsonify({"ok": True})
        d = request.get_json() or {}
        # 教师不能改 admin 任何字段
        if u["role"] == "admin" and request.user["role"] != "admin":
            return jsonify({"error": "无权修改管理员账号"}), 403
        if "name" in d: c.execute("UPDATE users SET name=? WHERE id=?", (d["name"], uid))
        if "role" in d:
            if request.user["role"] != "admin":
                # 教师改非 admin 用户的 role：值未变则允许；变了则 403
                if d["role"] != u["role"]:
                    return jsonify({"error": "只有管理员能修改角色"}), 403
                # role 值未变，跳过 UPDATE
            else:
                # admin 可以改任何人的 role；admin 改自己的 role 不允许
                if uid == request.user["id"] and d["role"] != u["role"]:
                    return jsonify({"error": "不能修改自己的角色"}), 400
                if d["role"] != u["role"]:
                    c.execute("UPDATE users SET role=? WHERE id=?", (d["role"], uid))
        if "class_id" in d: c.execute("UPDATE users SET class_id=? WHERE id=?", (d.get("class_id") or None, uid))
        if "student_no" in d: c.execute("UPDATE users SET student_no=? WHERE id=?", (d["student_no"], uid))
        if d.get("password"):
            # 改密码：作废该用户的所有 session（旧 token 立即失效）
            c.execute("UPDATE users SET password=? WHERE id=?", (hash_pw(d["password"]), uid))
            c.execute("DELETE FROM sessions WHERE user_id=?", (uid,))
        if "active" in d: c.execute("UPDATE users SET active=? WHERE id=?", (1 if d["active"] else 0, uid))
        return jsonify({"ok": True})

@app.route("/api/users/<int:uid>/change-password", methods=["POST"])
@login_required
def change_password(uid):
    """用户自己改密码：要求旧密码"""
    if uid != request.user["id"]:
        return jsonify({"error": "只能改自己的密码"}), 403
    d = request.get_json() or {}
    old_pwd = d.get("old_password") or ""
    new_pwd = d.get("new_password") or ""
    if not old_pwd or not new_pwd:
        return jsonify({"error": "旧密码和新密码必填"}), 400
    if len(new_pwd) < 6:
        return jsonify({"error": "新密码至少 6 位"}), 400
    with db() as c:
        u = c.execute("SELECT password FROM users WHERE id=?", (uid,)).fetchone()
        if not u: return jsonify({"error": "用户不存在"}), 404
        if u["password"] != hash_pw(old_pwd):
            return jsonify({"error": "旧密码错误"}), 401
        c.execute("UPDATE users SET password=? WHERE id=?", (hash_pw(new_pwd), uid))
        # 作废该用户所有 session
        c.execute("DELETE FROM sessions WHERE user_id=?", (uid,))
    return jsonify({"ok": True})


@app.route("/api/users/<int:uid>/reset-password", methods=["POST"])
@teacher_required
def reset_pwd(uid):
    d = request.get_json() or {}
    new_pwd = d.get("password", "123456")
    with db() as c:
        # 防越权：教师不能重置管理员密码
        target = c.execute("SELECT role FROM users WHERE id=?", (uid,)).fetchone()
        if not target: return jsonify({"error": "用户不存在"}), 404
        if target["role"] == "admin" and request.user["role"] != "admin":
            return jsonify({"error": "无权重置管理员密码"}), 403
        c.execute("UPDATE users SET password=? WHERE id=?", (hash_pw(new_pwd), uid))
        # 作废该用户所有 session
        c.execute("DELETE FROM sessions WHERE user_id=?", (uid,))
    return jsonify({"ok": True, "new_password": new_pwd})

# ---------- 知识点 ----------
@app.route("/api/knowledge", methods=["GET", "POST"])
@teacher_required
def knowledge():
    if request.method == "GET":
        with db() as c:
            rows = [dict(r) for r in c.execute("SELECT * FROM knowledge_points ORDER BY category, code").fetchall()]
        return jsonify(rows)
    d = request.get_json() or {}
    if not d.get("code") or not d.get("name"):
        return jsonify({"error": "编码和名称必填"}), 400
    with db() as c:
        try:
            cur = c.execute("INSERT INTO knowledge_points(code,name,category,description) VALUES(?,?,?,?)", (d["code"], d["name"], d.get("category", ""), d.get("description", "")))
            return jsonify({"ok": True, "id": cur.lastrowid})
        except Exception as e:
            if _is_duplicate_key_error(e):
                return jsonify({"error": "编码已存在"}), 400
            raise

@app.route("/api/knowledge/<int:kid>", methods=["PATCH", "DELETE"])
@teacher_required
def k_manage(kid):
    with db() as c:
        if request.method == "DELETE":
            c.execute("DELETE FROM knowledge_points WHERE id=?", (kid,))
            return jsonify({"ok": True})
        d = request.get_json() or {}
        if "name" in d: c.execute("UPDATE knowledge_points SET name=? WHERE id=?", (d["name"], kid))
        if "category" in d: c.execute("UPDATE knowledge_points SET category=? WHERE id=?", (d["category"], kid))
        if "description" in d: c.execute("UPDATE knowledge_points SET description=? WHERE id=?", (d["description"], kid))
        return jsonify({"ok": True})

# ---------- 题目 ----------
def _parse_q_options(opts):
    if not opts: return []
    if isinstance(opts, str):
        try: return json.loads(opts)
        except: return []
    return opts if isinstance(opts, list) else []

def _parse_q_answer(ans):
    if isinstance(ans, list): return ans
    if isinstance(ans, str):
        if ans.startswith("["):
            try: return json.loads(ans)
            except: pass
        return [a.strip() for a in re.split(r"[,，]", ans) if a.strip()] or [ans]
    return [str(ans)]

def _parse_q_kp(kp):
    if isinstance(kp, list): return kp
    if isinstance(kp, str):
        return [k.strip() for k in re.split(r"[;,]", kp) if k.strip()]
    return []

def _import_q(c, item, created_by=None):
    qtype = item.get("qtype")
    if not qtype: raise ValueError("题型必填")
    if not item.get("stem"): raise ValueError("题干不能为空")
    # 题组材料：stem 是材料文本，不需要 answer
    if qtype == "题组材料":
        opts, answer = [], []
    else:
        if not item.get("answer"): raise ValueError("答案必填")
        opts = _parse_q_options(item.get("options"))
        answer = _parse_q_answer(item.get("answer"))
        # 选择类题型必须有 ≥ 2 个选项
        if qtype in ("单选", "多选", "完形填空", "选词填空") and len(opts) < 2:
            raise ValueError("选择题至少需要 2 个选项")
    kp = _parse_q_kp(item.get("kp") or item.get("knowledge_points") or [])
    passage_id_raw = item.get("passage_id")
    passage_id = int(passage_id_raw) if (passage_id_raw is not None and str(passage_id_raw).strip() != "") else 0
    sub_idx = int(item.get("sub_idx") or 0)
    # unit_no：每个 unit 独立自增
    if passage_id <= 0:
        # 独立题或 passage 材料：分配新 unit_no
        max_no = c.execute("SELECT COALESCE(MAX(unit_no), 0) AS m FROM questions").fetchone()["m"]
        unit_no = max_no + 1
    else:
        # 子题：unit_no = parent 的 unit_no
        parent = c.execute("SELECT unit_no FROM questions WHERE id=?", (passage_id,)).fetchone()
        unit_no = parent["unit_no"] if parent else 0
    cur = c.execute("""INSERT INTO questions(source,year,region,qtype,difficulty,stem,options,answer,analysis,tags,kp,score,created_by,passage_id,sub_idx,unit_no)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (item.get("source"), item.get("year"), item.get("region"), qtype,
         int(item.get("difficulty") or 3), item["stem"],
         json.dumps(opts, ensure_ascii=False),
         json.dumps(answer, ensure_ascii=False),
         item.get("analysis", ""), item.get("tags", ""),
         json.dumps(kp, ensure_ascii=False),
         int(item.get("score") or 2), created_by, passage_id, sub_idx, unit_no))
    return cur.lastrowid

@app.route("/api/questions", methods=["GET"])
@login_required
def list_questions():
    qtype = request.args.get("qtype")
    kp = request.args.get("kp")
    difficulty = request.args.get("difficulty")
    keyword = request.args.get("q", "").strip()
    group_by_passage = request.args.get("group", "0") == "1"
    limit = min(int(request.args.get("limit", 500)), 5000)
    offset = int(request.args.get("offset", 0))
    with db() as c:
        sql = "SELECT * FROM questions WHERE active=1"
        args = []
        if qtype: sql += " AND qtype=?"; args.append(qtype)
        if difficulty: sql += " AND difficulty=?"; args.append(int(difficulty))
        if kp: sql += " AND kp LIKE ?"; args.append(f"%{kp}%")
        if keyword: sql += " AND (stem LIKE ? OR analysis LIKE ? OR tags LIKE ?)"; args.extend([f"%{keyword}%"]*3)
        # 排序：unit_no ASC, sub_idx ASC
        sql += " ORDER BY unit_no ASC, sub_idx ASC, id ASC"
        sql += " LIMIT ? OFFSET ?"; args.extend([limit, offset])
        rows = [dict(r) for r in c.execute(sql, args).fetchall()]
        for r in rows:
            r["options"] = json.loads(r["options"] or "[]")
            r["kp"] = json.loads(r["kp"] or "[]")
            if request.user["role"] == "student":
                r["answer"] = ""
            r["is_passage"] = (r["qtype"] == "题组材料")
            if r["passage_id"] == 0:
                # 独立题 或 passage 自身：unit_no 即题号
                r["display_no"] = str(r["unit_no"])
            else:
                # 子题：unit_no-sub_idx
                r["display_no"] = str(r["unit_no"]) + "-" + str(r["sub_idx"])
        if group_by_passage:
            # 分组返回：[{passage, children: []}] 或 [{question}]
            groups = []
            seen = {}
            for r in rows:
                if r["passage_id"] == 0:
                    groups.append({"type": "question", "data": r})
                else:
                    # 子题
                    if r["passage_id"] not in seen:
                        # 找 passage
                        pass_q = next((g["data"] for g in groups if g["type"] == "passage" and g["data"]["id"] == r["passage_id"]), None)
                        if pass_q is None:
                            pass_q_row = c.execute("SELECT * FROM questions WHERE id=?", (r["passage_id"],)).fetchone()
                            if pass_q_row:
                                pass_q = dict(pass_q_row)
                                pass_q["options"] = json.loads(pass_q["options"] or "[]")
                                pass_q["kp"] = json.loads(pass_q["kp"] or "[]")
                                if request.user["role"] == "student":
                                    pass_q["answer"] = ""
                                pass_q["display_no"] = str(passage_counter)  # 题组内统一编号
                                pass_q["is_passage"] = True
                                groups.append({"type": "passage", "data": pass_q})
                    grp = next((g for g in groups if g["type"] == "passage" and g["data"]["id"] == r["passage_id"]), None)
                    if grp:
                        grp.setdefault("children", []).append(r)
            return jsonify(groups)
        return jsonify(rows)

@app.route("/api/questions", methods=["POST"])
@teacher_required
def create_question():
    d = request.get_json() or {}
    try:
        with db() as c:
            qid = _import_q(c, d, request.user["id"])
        return jsonify({"ok": True, "id": qid})
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route("/api/questions/<int:qid>", methods=["GET", "PATCH", "DELETE"])
@teacher_required
def question_manage(qid):
    with db() as c:
        q = c.execute("SELECT * FROM questions WHERE id=?", (qid,)).fetchone()
        if not q: return jsonify({"error": "题目不存在"}), 404
        if request.method == "GET":
            if not q["active"]: return jsonify({"error": "题目已删除"}), 404
            d = dict(q)
            d["options"] = json.loads(d["options"] or "[]")
            d["kp"] = json.loads(d["kp"] or "[]")
            return jsonify(d)
        if request.method == "DELETE":
            c.execute("UPDATE questions SET active=0 WHERE id=?", (qid,))
            c.execute("DELETE FROM paper_questions WHERE question_id=?", (qid,))
            return jsonify({"ok": True})
        d = request.get_json() or {}
        cur = dict(q)
        for k in ("source","region","qtype","stem","analysis","tags"):
            if k in d: cur[k] = d[k]
        if "year" in d: cur["year"] = d["year"] or None
        if "difficulty" in d: cur["difficulty"] = int(d["difficulty"] or 3)
        if "score" in d: cur["score"] = int(d["score"] or 2)
        if "options" in d: cur["options"] = json.dumps(_parse_q_options(d["options"]), ensure_ascii=False)
        if "answer" in d: cur["answer"] = json.dumps(_parse_q_answer(d["answer"]), ensure_ascii=False)
        if "kp" in d: cur["kp"] = json.dumps(_parse_q_kp(d["kp"]), ensure_ascii=False)
        if "passage_id" in d: cur["passage_id"] = int(d["passage_id"] or 0)
        if "sub_idx" in d: cur["sub_idx"] = int(d["sub_idx"] or 0)
        c.execute("UPDATE questions SET source=?,year=?,region=?,qtype=?,difficulty=?,stem=?,options=?,answer=?,analysis=?,tags=?,kp=?,score=?,passage_id=?,sub_idx=? WHERE id=?",
            (cur["source"], cur["year"], cur["region"], cur["qtype"], cur["difficulty"], cur["stem"],
             cur["options"], cur["answer"], cur["analysis"], cur["tags"], cur["kp"], cur["score"],
             cur.get("passage_id", 0), cur.get("sub_idx", 0), qid))
        return jsonify({"ok": True})

def _make_excel_template():
    """生成 Excel 模板（3 个 sheet：题目数据 + 字段说明 + 题型规则）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="0ea5e9")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    # Sheet 1
    ws = wb.active
    ws.title = "题目数据"
    headers = ["qtype","stem","answer","option_a","option_b","option_c","option_d","option_e","option_f","kp","difficulty","score","year","source","analysis","passage_id","sub_idx"]
    ws.append(headers)
    for col in range(1, len(headers)+1):
        cell = ws.cell(1, col)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    samples = [
        # ─── 独立题 ───
        ["单选","She ___ to school every day.","B","go","goes","went","going","","","KP_GRAMMAR_TENSE",2,2,2023,"2023 北京中考","主语第三人称单数用 goes","",""],
        ["多选","Which are animals?","A,B,D","dog","cat","tree","bird","","","KP_VOCAB_NOUN",1,3,2023,"2023 上海中考","多个答案用英文逗号分隔","",""],
        ["翻译","Translate: 我喜欢学习英语。","I like studying English.","","","","","","","KP_TRANSLATION",3,5,2023,"2023 北京中考","","",""],
        # ─── 完形填空题组（passage_id=1，1 道材料 + 5 道子题）───
        ["题组材料","阅读下面短文,根据首字母、语境或所给单词的正确形式填空。\n\nTom is a 12-year-old boy. He 1___ (go) to school every day. He 2___ (like) English very much. His teacher says he is a 3___ (good) student. After school, Tom often 4___ (play) football with his friends. He 5___ (be) very happy.","","","","","","","","","KP_CLOZE",3,10,2023,"2023 北京中考","","",""],
        ["完形子题","Tom ___ to school every day.","B","go","goes","went","going","","","KP_GRAMMAR_TENSE",2,2,2023,"2023 北京中考","主语第三人称单数","1","1"],
        ["完形子题","He ___ English very much.","A","likes","like","is liking","liked","","","KP_GRAMMAR_TENSE",2,2,2023,"2023 北京中考","like doing/like to do","1","2"],
        ["完形子题","His teacher says he is a ___ student.","C","good","better","best","well","","","KP_VOCAB_ADJ",2,2,2023,"2023 北京中考","good 的最高级 best","1","3"],
        ["完形子题","After school, Tom often ___ football.","A","plays","play","is playing","played","","","KP_GRAMMAR_TENSE",2,2,2023,"2023 北京中考","often + 一般现在时","1","4"],
        ["完形子题","He ___ very happy.","A","is","are","be","am","","","KP_GRAMMAR_THERE_BE",2,2,2023,"2023 北京中考","He 系动词 is","1","5"],
        # ─── 阅读理解题组（passage_id=2，1 道材料 + 3 道子题）───
        ["题组材料","阅读下面短文,根据短文内容选择正确答案。\n\nMy name is Lisa. I am 13 years old. I study at No.5 Middle School. I usually get up at 6:30 in the morning. I have breakfast at 7:00. Then I go to school at 7:30. Classes begin at 8:00. I have lunch at school at 12:00. In the afternoon, classes finish at 4:30. I usually get home at 5:00.","","","","","","","","","KP_READING_DETAIL",3,15,2023,"2023 上海中考","","",""],
        ["阅读子题","How old is Lisa?","B","11","13","15","16","","","KP_READING_DETAIL",1,1,2023,"2023 上海中考","直接定位 13 years old","2","1"],
        ["阅读子题","What time does Lisa get up?","A","6:00","6:30","7:00","7:30","","","KP_READING_DETAIL",1,1,2023,"2023 上海中考","直接定位 6:30","2","2"],
        ["阅读子题","What time do classes begin?","C","7:00","7:30","8:00","12:00","","","KP_READING_DETAIL",1,1,2023,"2023 上海中考","直接定位 8:00","2","3"],
        # ─── 作文题组（passage_id=3, 1 道材料 + 1 道子题）───
        ["题组材料","根据以下要点,写一篇 80 词左右的英语短文,介绍你最好的朋友。\n要点:1. 他/她的姓名、年龄、外貌;2. 他/她的爱好;3. 你们常一起做的事情。","","","","","","","","","KP_WRITING",3,15,2023,"2023 北京中考","","",""],
        ["写作子题","My Best Friend","(主观题不计分)","","","","","","","","KP_WRITING",3,15,2023,"2023 北京中考","内容完整、语法正确即可","3","1"],
    ]
    for s in samples: ws.append(s)
    widths = [16, 50, 20, 15, 15, 15, 15, 15, 15, 22, 10, 8, 8, 18, 30, 10, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    # Sheet 2
    ws2 = wb.create_sheet("字段说明")
    ws2.append(["列名","是否必填","说明"])
    descs = [
        ["qtype","是","题型: 单选 / 多选 / 完形填空 / 选词填空 / 阅读理解 / 翻译 / 作文"],
        ["stem","是","题干内容"],
        ["answer","是","参考答案。单选填 A;多选/完形填 A,B,C(英文逗号)"],
        ["option_a","选择题必填","A 选项内容"],
        ["option_b","选择题必填","B 选项内容"],
        ["option_c","否","C 选项内容(可选)"],
        ["option_d","否","D 选项内容(可选)"],
        ["option_e","否","E 选项内容(可选)"],
        ["option_f","否","F 选项内容(可选)"],
        ["kp","否","知识点编码，多个用 ; 分隔，如 KP_GRAMMAR_TENSE;KP_VOCAB_NOUN"],
        ["difficulty","否","难度 1-5, 默认 3"],
        ["score","否","分值, 默认 2"],
        ["year","否","考试年份, 如 2023"],
        ["source","否","来源, 如 2023 北京中考"],
        ["analysis","否","答案解析"],
        ["passage_id","否","题组 id。0 = 独立题;同一题组的材料/子题填同一数字(如 1、2、3)"],
        ["sub_idx","否","子题在题组内的序号(1,2,3...)。材料题填 0,子题按顺序填"],
    ]
    for d in descs: ws2.append(d)
    for col in range(1, 4):
        c2 = ws2.cell(1, col)
        c2.fill = header_fill; c2.font = header_font
        c2.alignment = Alignment(horizontal="center")
    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 80
    # Sheet 3
    ws3 = wb.create_sheet("题型规则")
    ws3.append(["题型","最少选项","最多选项","答案格式"])
    rules = [
        ["单选",2,6,"填 1 个字母，如 B"],
        ["多选",2,6,"填多个字母逗号分隔，如 A,B,D"],
        ["完形填空",2,6,"填 1 个字母"],
        ["选词填空",2,6,"填 1 个字母"],
        ["阅读理解",2,6,"填 1 个字母"],
        ["翻译",0,0,"无需选项，answer 填译文"],
        ["作文",0,0,"无需选项，answer 填评分标准或留空"],
    ]
    for r in rules: ws3.append(r)
    for col in range(1, 5):
        c3 = ws3.cell(1, col)
        c3.fill = header_fill; c3.font = header_font
        c3.alignment = Alignment(horizontal="center")
    ws3.column_dimensions['A'].width = 14
    ws3.column_dimensions['B'].width = 12
    ws3.column_dimensions['C'].width = 12
    ws3.column_dimensions['D'].width = 30
    # Sheet 4: 题组规则（题号系统说明）
    ws4 = wb.create_sheet("题组规则")
    ws4.append(["题型组合","示例","题号系统","用法说明"])
    group_rules = [
        ["完形填空", "1 道材料 + 5 道子题", "材料 1;子题 1-1, 1-2, 1-3, 1-4, 1-5", "材料 qtype 填『题组材料』,sub_idx=0;子题 qtype 填『完形子题』,passage_id 填对应题组 id,sub_idx 填 1,2,3..."],
        ["阅读理解", "1 道材料 + N 道子题", "材料 2;子题 2-1, 2-2, ...", "材料 qtype=题组材料,子题 qtype=阅读子题,共享 passage_id"],
        ["写作", "1 道材料 + 1 道子题(学生写作处)", "材料 3;子题 3-1", "材料 qtype=题组材料,子题 qtype=写作子题,共享 passage_id"],
        ["独立题", "1 道题", "1, 2, 3, ...", "passage_id 留空或 0,sub_idx 留空或 0"],
    ]
    for r in group_rules: ws4.append(r)
    for col in range(1, 5):
        c2 = ws4.cell(1, col)
        c2.fill = header_fill; c2.font = header_font
        c2.alignment = Alignment(horizontal="center")
    ws4.column_dimensions['A'].width = 16
    ws4.column_dimensions['B'].width = 28
    ws4.column_dimensions['C'].width = 32
    ws4.column_dimensions['D'].width = 60
    import io as _io
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

@app.route("/api/questions/import", methods=["GET", "POST"])
@teacher_required
def import_q():
    if request.method == "GET" and request.args.get("format") == "template":
        buf = _make_excel_template()
        from flask import send_file
        return send_file(buf, as_attachment=True, download_name="题库导入模板.xlsx",
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    if "file" not in request.files: return jsonify({"error": "未上传文件"}), 400
    f = request.files["file"]
    name = f.filename.lower()
    if name.endswith(".json"):
        data = json.load(f)
        if isinstance(data, dict) and "questions" in data: data = data["questions"]
        ok, err = 0, []
        with db() as c:
            for i, item in enumerate(data):
                try: _import_q(c, item, request.user["id"]); ok += 1
                except Exception as e: err.append({"row": i+1, "error": str(e)})
        return jsonify({"inserted": ok, "failed": len(err), "errors": err[:10]})
    if name.endswith((".xlsx", ".xlsm")):
        try:
            from openpyxl import load_workbook
        except ImportError:
            return jsonify({"error": "需要 openpyxl: pip install openpyxl"}), 500
        wb = load_workbook(f, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(c).strip() if c else "" for c in rows[0]]
        # 两遍扫描：第一遍找到题组材料行（按行号），第二遍为子题分配 passage_id
        # 如果用户已经填了 passage_id 且有效（数据库中存在该 id），保留；否则用占位符
        # 简化：所有题按顺序入库，passage_id 留空或 0 的子题归到上一个题组材料
        current_passage_id = 0
        current_passage_row = 0
        passage_seq = {}  # passage_id 临时分配：1, 2, 3...  按出现顺序
        # 收集现有 passage (qtype=题组材料) 的最大 unit_no
        max_existing_unit = 0
        try:
            max_existing_unit = c.execute("SELECT COALESCE(MAX(unit_no), 0) AS m FROM questions").fetchone()["m"]
        except: pass
        ok, err = 0, []
        with db() as c:
            for i, row in enumerate(rows[1:], start=2):
                if not row or not any(row): continue
                item = dict(zip(headers, row))
                qtype_str = str(item.get("qtype", "")).strip()
                # 自动分配 passage_id
                if qtype_str == "题组材料":
                    max_existing_unit += 1
                    current_passage_id = -max_existing_unit  # 用负数占位（避免与数据库 id 冲突）
                    current_passage_row = i
                elif qtype_str in ("完形子题", "阅读子题", "写作子题"):
                    if current_passage_id < 0:
                        # 子题归到上一个 passage
                        item["passage_id"] = current_passage_id
                    else:
                        # 没有 passage 在前，子题变成独立题
                        item["passage_id"] = 0
                else:
                    # 普通题：passage_id=0
                    item["passage_id"] = 0
                opts = []
                for letter in ["A","B","C","D","E","F"]:
                    v = item.pop(f"option_{letter.lower()}", None)
                    if v: opts.append({"label": letter, "content": str(v)})
                item["options"] = opts
                item["kp"] = [k.strip() for k in str(item.get("kp","") or "").split(";") if k.strip()]
                # 安全转 int 字段
                for int_key in ("year", "difficulty", "score", "passage_id", "sub_idx"):
                    v = item.get(int_key)
                    if v is not None and str(v).strip() != "":
                        try: item[int_key] = int(float(str(v).strip()))
                        except: item.pop(int_key, None)
                    else:
                        item.pop(int_key, None)
                try: _import_q(c, item, request.user["id"]); ok += 1
                except Exception as e: err.append({"row": i, "error": str(e)})
        return jsonify({"inserted": ok, "failed": len(err), "errors": err[:10]})
    if name.endswith(".csv"):
        content = f.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))
        ok, err = 0, []
        with db() as c:
            for i, item in enumerate(reader, start=2):
                # 合并 option_a~f 为 options 数组
                opts = []
                for k in ['option_a','option_b','option_c','option_d','option_e','option_f']:
                    v = (item.get(k) or '').strip()
                    if v: opts.append({"label": k[-1].upper(), "content": v})
                item['options'] = opts
                try: _import_q(c, item, request.user["id"]); ok += 1
                except Exception as e: err.append({"row": i, "error": str(e)})
        return jsonify({"inserted": ok, "failed": len(err), "errors": err[:10]})
    return jsonify({"error": "支持 .json/.xlsx/.csv"}), 400

@app.route("/api/questions/summary")
@login_required
def q_summary():
    with db() as c:
        total = c.execute("SELECT COUNT(*) c FROM questions WHERE active=1").fetchone()["c"]
        by_type = [dict(r) for r in c.execute("SELECT qtype name, COUNT(*) count FROM questions WHERE active=1 GROUP BY qtype").fetchall()]
        kp_count = {}
        for r in c.execute("SELECT kp FROM questions WHERE active=1 AND kp IS NOT NULL AND kp != '[]' and kp != 'null' and kp != ''").fetchall():
            try:
                for k in json.loads(r["kp"] or "[]"):
                    kp_count[k] = kp_count.get(k, 0) + 1
            except: pass
        return jsonify({"total": total, "by_type": by_type, "by_kp": [{"name": k, "count": v} for k, v in sorted(kp_count.items(), key=lambda x: -x[1])]})

# ---------- 试卷 ----------
@app.route("/api/papers", methods=["GET", "POST"])
@login_required
def papers():
    if request.method == "GET":
        with db() as c:
            rows = [dict(r) for r in c.execute("SELECT p.*, c.name AS class_name FROM papers p LEFT JOIN classes c ON c.id=p.class_id ORDER BY p.id DESC").fetchall()]
            for r in rows:
                r["question_count"] = c.execute("SELECT COUNT(*) c FROM paper_questions WHERE paper_id=?", (r["id"],)).fetchone()["c"]
        if request.user["role"] == "student":
            rows = [r for r in rows if r["published"] and (not r["class_id"] or r["class_id"] == request.user["class_id"])]
        return jsonify(rows)
    if request.user["role"] not in ("admin", "teacher"): return jsonify({"error": "无权限"}), 403
    d = request.get_json() or {}
    with db() as c:
        cur = c.execute("""INSERT INTO papers(name,description,duration_minutes,total_score,class_id,published,created_by) VALUES(?,?,?,?,?,?,?)""",
            (d["name"], d.get("description",""), int(d.get("duration_minutes",120)), int(d.get("total_score",100)), d.get("class_id") or None, 1 if d.get("published", True) else 0, request.user["id"]))
        pid = cur.lastrowid
        for idx, qid in enumerate(d.get("question_ids", [])):
            q = c.execute("SELECT score FROM questions WHERE id=?", (qid,)).fetchone()
            if not q: continue
            if USE_POSTGRES:
                c.execute("INSERT INTO paper_questions(paper_id,question_id,order_index,score) VALUES(?,?,?,?) ON CONFLICT (paper_id, question_id) DO NOTHING", (pid, qid, idx, q["score"]))
            else:
                c.execute("INSERT OR IGNORE INTO paper_questions(paper_id,question_id,order_index,score) VALUES(?,?,?,?)", (pid, qid, idx, q["score"]))
        c.execute("UPDATE papers SET question_count=? WHERE id=?", (len(d.get("question_ids", [])), pid))
        return jsonify({"ok": True, "id": pid})

@app.route("/api/papers/<int:pid>", methods=["GET", "DELETE"])
@login_required
def paper_manage(pid):
    with db() as c:
        p = c.execute("SELECT * FROM papers WHERE id=?", (pid,)).fetchone()
        if not p: return jsonify({"error": "试卷不存在"}), 404
        if request.method == "GET":
            d = dict(p)
            qs = [dict(r) for r in c.execute("""SELECT q.*, pq.order_index, pq.score AS paper_score FROM paper_questions pq JOIN questions q ON q.id=pq.question_id WHERE pq.paper_id=? AND q.active=1 ORDER BY pq.order_index""", (pid,)).fetchall()]
            for q in qs:
                q["options"] = json.loads(q["options"] or "[]")
                q["kp"] = json.loads(q["kp"] or "[]")
                if request.user["role"] == "student": q["answer"] = ""
            d["questions"] = qs
            return jsonify(d)
        if request.user["role"] not in ("admin", "teacher"): return jsonify({"error": "无权限"}), 403
        c.execute("DELETE FROM paper_questions WHERE paper_id=?", (pid,))
        c.execute("DELETE FROM papers WHERE id=?", (pid,))
    return jsonify({"ok": True})

@app.route("/api/papers/<int:pid>/update", methods=["POST"])
@login_required
def paper_update(pid):
    if request.user["role"] not in ("admin", "teacher"): return jsonify({"error": "无权限"}), 403
    d = request.get_json() or {}
    with db() as c:
        if "name" in d: c.execute("UPDATE papers SET name=? WHERE id=?", (d["name"], pid))
        if "description" in d: c.execute("UPDATE papers SET description=? WHERE id=?", (d["description"], pid))
        if "duration_minutes" in d: c.execute("UPDATE papers SET duration_minutes=? WHERE id=?", (int(d["duration_minutes"]), pid))
        if "total_score" in d: c.execute("UPDATE papers SET total_score=? WHERE id=?", (int(d["total_score"]), pid))
        if "class_id" in d: c.execute("UPDATE papers SET class_id=? WHERE id=?", (d.get("class_id") or None, pid))
        if "published" in d: c.execute("UPDATE papers SET published=? WHERE id=?", (1 if d["published"] else 0, pid))
        if "question_ids" in d:
            c.execute("DELETE FROM paper_questions WHERE paper_id=?", (pid,))
            for idx, qid in enumerate(d["question_ids"]):
                q = c.execute("SELECT score FROM questions WHERE id=?", (qid,)).fetchone()
                if not q: continue
                c.execute("INSERT INTO paper_questions(paper_id,question_id,order_index,score) VALUES(?,?,?,?)", (pid, qid, idx, q["score"]))
            c.execute("UPDATE papers SET question_count=? WHERE id=?", (len(d["question_ids"]), pid))
    return jsonify({"ok": True})

# ---------- 答题 ----------
def is_correct(q, user_ans):
    try: correct = json.loads(q["answer"])
    except: correct = [q["answer"]]
    if not user_ans: return False, 0
    qt = q["qtype"]
    if qt == "单选":
        ok = user_ans.strip().upper() == (correct[0] if correct else "").strip().upper()
    elif qt in ("多选", "完形填空", "选词填空"):
        ua = sorted([a.strip().upper() for a in re.split(r"[,，]", user_ans) if a.strip()])
        ca = sorted([str(a).strip().upper() for a in correct])
        ok = ua == ca and len(ua) > 0
    else:
        ok = False
    return ok, (q["score"] if ok else 0)

@app.route("/api/attempts/start", methods=["POST"])
@login_required
def start_attempt():
    d = request.get_json() or {}
    mode = d.get("mode", "free")
    count = min(int(d.get("count", 10)), 50)
    paper_id = d.get("paper_id")
    qids = []
    paper_info = None
    with db() as c:
        if mode == "exam" and paper_id:
            p = c.execute("SELECT * FROM papers WHERE id=?", (paper_id,)).fetchone()
            if not p: return jsonify({"error": "试卷不存在"}), 404
            paper_info = dict(p)
            qids = [r["question_id"] for r in c.execute("SELECT question_id FROM paper_questions WHERE paper_id=? AND question_id IN (SELECT id FROM questions WHERE active=1) ORDER BY order_index", (paper_id,)).fetchall()]
        elif mode == "review":
            qids = [r["question_id"] for r in c.execute("SELECT question_id FROM wrong_book WHERE user_id=? AND mastered=0 ORDER BY last_wrong_at DESC LIMIT ?", (request.user["id"], count)).fetchall()]
            if len(qids) < count:
                ex_ids = ",".join("?"*len(qids)) if qids else "0"
                extra = [r["id"] for r in c.execute(f"SELECT id FROM questions WHERE active=1 AND id NOT IN ({ex_ids}) ORDER BY RANDOM() LIMIT ?", qids + [count - len(qids)]).fetchall()]
                qids += extra
        else:
            qtype = d.get("qtype")
            if qtype:
                qids = [r["id"] for r in c.execute("SELECT id FROM questions WHERE active=1 AND qtype=? ORDER BY RANDOM() LIMIT ?", (qtype, count)).fetchall()]
            else:
                qids = [r["id"] for r in c.execute("SELECT id FROM questions WHERE active=1 ORDER BY RANDOM() LIMIT ?", (count,)).fetchall()]
        if not qids: return jsonify({"error": "题库为空，请联系老师导入题目"}), 400
        duration = paper_info["duration_minutes"] * 60 if paper_info else None
        cur = c.execute("INSERT INTO attempts(user_id,mode,paper_id,status,qcount,qids,duration_seconds) VALUES(?,?,?,?,?,?,?)",
            (request.user["id"], mode, paper_id, "in_progress", len(qids), json.dumps(qids), duration))
        aid = cur.lastrowid
        ph = ",".join("?"*len(qids))
        qs = [dict(r) for r in c.execute(f"SELECT * FROM questions WHERE id IN ({ph}) AND active=1", qids).fetchall()]
        qmap = {q["id"]: q for q in qs}
        questions = []
        for qid in qids:
            q = qmap.get(qid)
            if not q: continue
            questions.append({
                "id": q["id"], "qtype": q["qtype"], "difficulty": q["difficulty"], "stem": q["stem"],
                "score": q["score"], "source": q["source"], "year": q["year"],
                "options": json.loads(q["options"] or "[]"), "kp": json.loads(q["kp"] or "[]")
            })
    return jsonify({"attempt_id": aid, "mode": mode, "paper_id": paper_id, "paper_info": paper_info, "questions": questions, "count": len(questions)})

@app.route("/api/attempts/answer", methods=["POST"])
@login_required
def submit_answer():
    d = request.get_json() or {}
    aid, qid = int(d["attempt_id"]), int(d["question_id"])
    ua = (d.get("user_answer") or "").strip()
    time_spent = int(d.get("time_spent", 0))
    with db() as c:
        a = c.execute("SELECT * FROM attempts WHERE id=? AND user_id=?", (aid, request.user["id"])).fetchone()
        if not a: return jsonify({"error": "会话不存在"}), 404
        if a["status"] != "in_progress": return jsonify({"error": "会话已结束"}), 400
        is_review_mode = (a["mode"] == "review")
        q = c.execute("SELECT * FROM questions WHERE id=?", (qid,)).fetchone()
        if not q: return jsonify({"error": "题目不存在"}), 404
        # 检查是否已答过该题（同 attempt 内），若是则覆盖（回滚错题本旧状态）
        prev = c.execute("SELECT * FROM answers WHERE attempt_id=? AND question_id=?", (aid, qid)).fetchone()
        if prev:
            # 回滚错题本：之前答错，现在若答对，把错题本标为已掌握
            if not prev["is_correct"]:
                # 之前是错的，现在新答案若是正确，把 wrong_count 减 1
                # 但只在错题本中确实存在该记录时操作
                pass  # 下面统一处理
        correct, score = is_correct(q, ua)
        # 错题本处理：答对时 -1，减到 0 自动删除
        if correct:
            ex_for_dec = c.execute("SELECT * FROM wrong_book WHERE user_id=? AND question_id=?", (request.user["id"], qid)).fetchone()
            if ex_for_dec:
                # 跳过 review 模式下 prev=空（第一次答对）的本就是错题
                if prev is None or not prev["is_correct"]:
                    c.execute("UPDATE wrong_book SET wrong_count = wrong_count - 1 WHERE user_id=? AND question_id=?", (request.user["id"], qid))
                    c.execute("DELETE FROM wrong_book WHERE user_id=? AND question_id=? AND wrong_count <= 0", (request.user["id"], qid))
        if not correct:
            ex = c.execute("SELECT * FROM wrong_book WHERE user_id=? AND question_id=?", (request.user["id"], qid)).fetchone()
            if ex:
                if is_review_mode:
                    # review 模式答错：不增加 wrong_count（已经是错题了），只更新时间和取消 mastered
                    c.execute("UPDATE wrong_book SET last_wrong_at=CURRENT_TIMESTAMP, mastered=0 WHERE id=?", (ex["id"],))
                elif prev is None or prev["is_correct"]:
                    # 新 attempt 答错 或 同 attempt 内对→错：+1
                    c.execute("UPDATE wrong_book SET wrong_count=wrong_count+1, last_wrong_at=CURRENT_TIMESTAMP, mastered=0 WHERE id=?", (ex["id"],))
                else:
                    # 同 attempt 内又错：只更新 last_wrong_at
                    c.execute("UPDATE wrong_book SET last_wrong_at=CURRENT_TIMESTAMP, mastered=0 WHERE id=?", (ex["id"],))
            else:
                c.execute("INSERT INTO wrong_book(user_id,question_id) VALUES(?,?)", (request.user["id"], qid))
        # 写入/更新 answers
        if prev:
            c.execute("UPDATE answers SET user_answer=?, is_correct=?, score_earned=?, time_spent=?, answered_at=CURRENT_TIMESTAMP WHERE id=?", (ua, 1 if correct else 0, score, time_spent, prev["id"]))
        else:
            c.execute("INSERT INTO answers(attempt_id,user_id,question_id,user_answer,is_correct,score_earned,time_spent) VALUES(?,?,?,?,?,?,?)",
                (aid, request.user["id"], qid, ua, 1 if correct else 0, score, time_spent))
    return jsonify({"correct": correct, "correct_answer": q["answer"], "analysis": q["analysis"], "score_earned": score})

@app.route("/api/attempts/<int:aid>/submit", methods=["POST"])
@login_required
def submit_attempt(aid):
    with db() as c:
        a = c.execute("SELECT * FROM attempts WHERE id=? AND user_id=?", (aid, request.user["id"])).fetchone()
        if not a: return jsonify({"error": "会话不存在"}), 404
        if a["paper_id"]:
            p = c.execute("SELECT id FROM papers WHERE id=?", (a["paper_id"],)).fetchone()
            if not p: return jsonify({"error": "试卷已被删除，无法提交"}), 400
        if a["status"] == "submitted": return jsonify({"ok": True, "attempt": dict(a)})
        ans = [dict(r) for r in c.execute("SELECT * FROM answers WHERE attempt_id=?", (aid,)).fetchall()]
        total_score = sum(r["score_earned"] for r in ans)
        correct_count = sum(1 for r in ans if r["is_correct"])
        qids = json.loads(a["qids"])
        row = c.execute("SELECT SUM(score) s FROM questions WHERE id IN (" + ",".join("?"*len(qids)) + ")", qids).fetchone()
        max_score = row["s"] or 0
        duration = None
        if a["started_at"]:
            try:
                started = datetime.strptime(a["started_at"], "%Y-%m-%d %H:%M:%S")
                duration = int((datetime.utcnow() - started).total_seconds())
            except: duration = None
        c.execute("UPDATE attempts SET status='submitted', correct=?, score_earned=?, max_score=?, submitted_at=CURRENT_TIMESTAMP, duration_seconds=? WHERE id=?",
            (correct_count, total_score, max_score, duration, aid))
        a = c.execute("SELECT * FROM attempts WHERE id=?", (aid,)).fetchone()
        if a["paper_id"] and a["qids"]:
            qids2 = json.loads(a["qids"])
            for r in ans:
                if not r["is_correct"]:
                    ex = c.execute("SELECT * FROM wrong_book WHERE user_id=? AND question_id=?", (request.user["id"], r["question_id"])).fetchone()
                    if ex:
                        c.execute("UPDATE wrong_book SET wrong_count=wrong_count+1, last_wrong_at=CURRENT_TIMESTAMP, mastered=0 WHERE id=?", (ex["id"],))
                    else:
                        c.execute("INSERT INTO wrong_book(user_id,question_id) VALUES(?,?)", (request.user["id"], r["question_id"]))
    return jsonify({"ok": True, "attempt": dict(a), "total": total_score, "max_score": max_score, "correct": correct_count})

@app.route("/api/attempts/history")
@login_required
def my_history():
    with db() as c:
        rows = [dict(r) for r in c.execute("""SELECT a.*, p.name AS paper_name FROM attempts a LEFT JOIN papers p ON p.id=a.paper_id WHERE a.user_id=? AND a.status="submitted" ORDER BY a.id DESC LIMIT 30""", (request.user["id"],)).fetchall()]
    return jsonify(rows)

@app.route("/api/attempts/<int:aid>/review")
@login_required
def review_attempt(aid):
    with db() as c:
        a = c.execute("SELECT * FROM attempts WHERE id=? AND user_id=?", (aid, request.user["id"])).fetchone()
        if not a: return jsonify({"error": "会话不存在"}), 404
        ans = [dict(r) for r in c.execute("SELECT * FROM answers WHERE attempt_id=?", (aid,)).fetchall()]
        qids = [r["question_id"] for r in ans]
        qs = [dict(r) for r in c.execute("SELECT * FROM questions WHERE id IN (" + ",".join("?"*len(qids)) + ")", qids).fetchall()] if qids else []
        qmap = {q["id"]: q for q in qs}
        for r in ans:
            q = qmap.get(r["question_id"])
            if q:
                q["options"] = json.loads(q["options"] or "[]")
                r["question"] = q
        return jsonify({"attempt": dict(a), "answers": ans})

# ---------- 学生统计 ----------
@app.route("/api/stats/me")
@login_required
def my_stats():
    uid = request.user["id"]
    with db() as c:
        total_q = c.execute("SELECT COUNT(*) c FROM answers WHERE user_id=?", (uid,)).fetchone()["c"]
        total_c = c.execute("SELECT COUNT(*) c FROM answers WHERE user_id=? AND is_correct=1", (uid,)).fetchone()["c"]
        attempts_n = c.execute("SELECT COUNT(*) c FROM attempts WHERE user_id=? AND status='submitted'", (uid,)).fetchone()["c"]
        accuracy = total_c / total_q if total_q else 0
        rows = c.execute("SELECT date(answered_at) d, COUNT(*) q, SUM(is_correct) c FROM answers WHERE user_id=? GROUP BY d ORDER BY d DESC LIMIT 14", (uid,)).fetchall()
        trend = [{"date": r["d"], "qcount": r["q"], "correct": r["c"], "accuracy": (r["c"]/r["q"]) if r["q"] else 0} for r in reversed(rows)]
        kp_stats = {}
        for r in c.execute("SELECT q.kp, a.is_correct FROM answers a JOIN questions q ON q.id=a.question_id WHERE a.user_id=?", (uid,)).fetchall():
            try: kps = json.loads(r["kp"] or "[]")
            except: kps = []
            for k in kps:
                kp_stats.setdefault(k, {"kp": k, "attempted":0, "correct":0})
                kp_stats[k]["attempted"] += 1
                if r["is_correct"]: kp_stats[k]["correct"] += 1
        # 把 kp 编码映射成名称
        kp_map = {k["code"]: k["name"] for k in c.execute("SELECT code, name FROM knowledge_points").fetchall()}
        for code, v in kp_stats.items():
            v["kp_name"] = kp_map.get(code, code)
        kp_list = []
        for v in kp_stats.values():
            v["accuracy"] = v["correct"]/v["attempted"] if v["attempted"] else 0
            kp_list.append(v)
        kp_list.sort(key=lambda x: -x["attempted"])
        type_stats = [dict(r) for r in c.execute("""SELECT q.qtype, COUNT(*) q, SUM(a.is_correct) c FROM answers a JOIN questions q ON q.id=a.question_id WHERE a.user_id=? GROUP BY q.qtype""", (uid,)).fetchall()]
        for t in type_stats: t["accuracy"] = t["c"]/t["q"] if t["q"] else 0
        wrong = [dict(r) for r in c.execute("""SELECT w.*, q.stem, q.qtype, q.source, q.year FROM wrong_book w JOIN questions q ON q.id=w.question_id WHERE w.user_id=? AND w.mastered=0 ORDER BY w.last_wrong_at DESC LIMIT 50""", (uid,)).fetchall()]
        papers = [dict(r) for r in c.execute("""SELECT a.id AS attempt_id, a.score_earned, a.max_score, a.correct, a.qcount, a.submitted_at, p.name AS paper_name, p.id AS paper_id FROM attempts a JOIN papers p ON p.id=a.paper_id WHERE a.user_id=? AND a.mode='exam' AND a.status='submitted' ORDER BY a.id DESC LIMIT 20""", (uid,)).fetchall()]
    return jsonify({"total_q": total_q, "total_c": total_c, "accuracy": round(accuracy,4), "attempts": attempts_n, "trend": trend, "kp_stats": kp_list, "type_stats": type_stats, "wrong_book": wrong, "paper_scores": papers})

@app.route("/api/wrong-book")
@login_required
def get_wrong():
    with_mastered = request.args.get("mastered") == "true"
    sql = "SELECT w.*, q.stem, q.qtype, q.source, q.year FROM wrong_book w JOIN questions q ON q.id=w.question_id WHERE w.user_id=?"
    args = [request.user["id"]]
    if not with_mastered:
        sql += " AND w.mastered=0"
    sql += " ORDER BY w.last_wrong_at DESC"
    with db() as c:
        rows = [dict(r) for r in c.execute(sql, args).fetchall()]
    return jsonify(rows)

@app.route("/api/wrong-book/master/<int:wid>", methods=["POST"])
@login_required
def mark_mastered(wid):
    with db() as c:
        c.execute("UPDATE wrong_book SET mastered=1 WHERE id=? AND user_id=?", (wid, request.user["id"]))
    return jsonify({"ok": True})

@app.route("/api/recommend")
@login_required
def recommend():
    uid = request.user["id"]
    count = min(max(int(request.args.get("count", 15)), 0), 50)
    chosen = []; used = set(); rationale = []
    with db() as c:
        # 题库总量检查
        total_q = c.execute("SELECT COUNT(*) c FROM questions WHERE active=1").fetchone()["c"]
        if total_q == 0 or count == 0:
            return jsonify({"qids": [], "rationale": "题库暂无题目，请老师先导入题目" if total_q == 0 else "推荐数量为 0"})
        wrong = [r["question_id"] for r in c.execute("SELECT question_id FROM wrong_book WHERE user_id=? AND mastered=0 ORDER BY last_wrong_at DESC LIMIT ?", (uid, count)).fetchall()]
        for q in wrong:
            if q not in used: chosen.append(q); used.add(q)
        if wrong: rationale.append(f"含 {len(wrong)} 道近期错题")
        weak = []
        for r in c.execute("SELECT q.kp, a.is_correct FROM answers a JOIN questions q ON q.id=a.question_id WHERE a.user_id=?", (uid,)).fetchall():
            try: kps = json.loads(r["kp"] or "[]")
            except: kps = []
            for k in kps: weak.append((k, r["is_correct"]))
        kp_acc = {}
        for k, c_v in weak:
            kp_acc.setdefault(k, [0, 0])
            kp_acc[k][0] += 1
            if c_v: kp_acc[k][1] += 1
        weak_top = sorted([(k, v[1]/v[0]) for k, v in kp_acc.items() if v[0] >= 2], key=lambda x: x[1])[:3]
        for kp_code, acc in weak_top:
            kp_qs = [r["id"] for r in c.execute("SELECT id FROM questions WHERE active=1 AND kp LIKE ?", (f"%{kp_code}%",)).fetchall()]
            random.shuffle(kp_qs)
            for q in kp_qs:
                if q not in used and len(chosen) < count: chosen.append(q); used.add(q)
        if weak_top: rationale.append("聚焦弱项: " + ", ".join(k for k, _ in weak_top))
        while len(chosen) < count:
            if used:
                ph = ",".join("?"*len(used))
                row = c.execute(f"SELECT id FROM questions WHERE active=1 AND id NOT IN ({ph}) ORDER BY RANDOM() LIMIT 1", list(used)).fetchone()
            else:
                row = c.execute("SELECT id FROM questions WHERE active=1 ORDER BY RANDOM() LIMIT 1").fetchone()
            if not row: break
            chosen.append(row["id"]); used.add(row["id"])
    random.shuffle(chosen)
    chosen = chosen[:count]
    rationale_text = "推荐理由:\n" + "\n".join("• " + r for r in rationale) if rationale else "随机推荐"
    return jsonify({"qids": chosen, "rationale": rationale_text})

# ---------- 教师统计 ----------
@app.route("/api/stats/teacher")
@teacher_required
def teacher_stats():
    class_id = request.args.get("class_id")
    with db() as c:
        classes = [dict(r) for r in c.execute("SELECT * FROM classes ORDER BY id").fetchall()]
        total_students = c.execute("SELECT COUNT(*) c FROM users WHERE role='student' AND active=1").fetchone()["c"]
        total_questions = c.execute("SELECT COUNT(*) c FROM questions WHERE active=1").fetchone()["c"]
        total_papers = c.execute("SELECT COUNT(*) c FROM papers").fetchone()["c"]
        total_attempts = c.execute("SELECT COUNT(*) c FROM attempts WHERE status='submitted'").fetchone()["c"]
        students_sql = "SELECT u.id, u.name, u.username, u.class_id, c.name AS class_name, (SELECT COUNT(*) FROM answers WHERE user_id=u.id) AS total_q, (SELECT COUNT(*) FROM answers WHERE user_id=u.id AND is_correct=1) AS total_c, (SELECT COUNT(*) FROM attempts WHERE user_id=u.id AND status='submitted') AS attempts, (SELECT AVG(score_earned*1.0/NULLIF(max_score,0)) FROM attempts WHERE user_id=u.id AND max_score>0 AND status='submitted') AS avg_ratio FROM users u LEFT JOIN classes c ON c.id=u.class_id WHERE u.role='student' AND u.active=1"
        args = []
        if class_id: students_sql += " AND u.class_id=?"; args.append(class_id)
        students_sql += " ORDER BY c.id, u.id"
        students = [dict(r) for r in c.execute(students_sql, args).fetchall()]
        for s in students:
            s["accuracy"] = s["total_c"]/s["total_q"] if s["total_q"] else 0
            s["avg_score_ratio"] = s["avg_score_ratio"] if "avg_score_ratio" in s else 0
            s["avg_score_ratio"] = s.get("avg_ratio", 0) or 0
        paper_stats = [dict(r) for r in c.execute("""SELECT p.id, p.name, p.total_score, p.class_id, c.name AS class_name, COUNT(a.id) AS submissions, AVG(a.score_earned) avg_score, MAX(a.score_earned) max_score, MIN(a.score_earned) min_score FROM papers p LEFT JOIN attempts a ON a.paper_id=p.id AND a.status='submitted' LEFT JOIN classes c ON c.id=p.class_id GROUP BY p.id ORDER BY p.id DESC""").fetchall()]
        kp_total = {}
        for r in c.execute("SELECT q.kp, a.is_correct FROM answers a JOIN questions q ON q.id=a.question_id").fetchall():
            try: kps = json.loads(r["kp"] or "[]")
            except: kps = []
            for k in kps:
                kp_total.setdefault(k, [0, 0])
                kp_total[k][0] += 1
                if r["is_correct"]: kp_total[k][1] += 1
        kp_map = {k["code"]: k["name"] for k in c.execute("SELECT code, name FROM knowledge_points").fetchall()}
        kp_summary = [{"kp": k, "kp_name": kp_map.get(k, k), "attempted": v[0], "correct": v[1], "accuracy": v[1]/v[0]} for k, v in kp_total.items()]
        kp_summary.sort(key=lambda x: x["accuracy"])
        type_dist = [dict(r) for r in c.execute("SELECT qtype name, COUNT(*) count FROM questions WHERE active=1 GROUP BY qtype").fetchall()]
    return jsonify({"classes": classes, "total_students": total_students, "total_questions": total_questions, "total_papers": total_papers, "total_attempts": total_attempts, "students": students, "paper_stats": paper_stats, "kp_summary": kp_summary, "type_dist": type_dist})

@app.route("/api/stats/paper/<int:pid>")
@teacher_required
def paper_detail_stats(pid):
    with db() as c:
        p = c.execute("SELECT * FROM papers WHERE id=?", (pid,)).fetchone()
        if not p: return jsonify({"error": "试卷不存在"}), 404
        q_stats = [dict(r) for r in c.execute("""SELECT q.id, q.qtype, q.stem, pq.score, (SELECT COUNT(*) FROM answers a JOIN attempts at ON at.id=a.attempt_id WHERE a.question_id=q.id AND at.paper_id=?) AS attempts, (SELECT COUNT(*) FROM answers a JOIN attempts at ON at.id=a.attempt_id WHERE a.question_id=q.id AND at.paper_id=? AND a.is_correct=1) AS correct FROM paper_questions pq JOIN questions q ON q.id=pq.question_id WHERE pq.paper_id=? ORDER BY pq.order_index""", (pid, pid, pid)).fetchall()]
        for q in q_stats:
            q["accuracy"] = q["correct"]/q["attempts"] if q["attempts"] else 0
            q["stem"] = q["stem"][:60] + ("..." if len(q["stem"]) > 60 else "")
        scores = [dict(r) for r in c.execute("SELECT score_earned, max_score, user_id FROM attempts WHERE paper_id=? AND status='submitted' AND max_score>0", (pid,)).fetchall()]
        ranges = {"90-100":0, "80-89":0, "70-79":0, "60-69":0, "0-59":0}
        for s in scores:
            r = (s["score_earned"]/s["max_score"]*100) if s["max_score"] else 0
            if r >=90: ranges["90-100"] += 1
            elif r >=80: ranges["80-89"] += 1
            elif r >=70: ranges["70-79"] += 1
            elif r >=60: ranges["60-69"] += 1
            else: ranges["0-59"] += 1
        rows = [dict(r) for r in c.execute("""SELECT u.id, u.name, u.username, c.name AS class_name, a.score_earned, a.max_score, a.correct, a.qcount, a.submitted_at FROM attempts a JOIN users u ON u.id=a.user_id LEFT JOIN classes c ON c.id=u.class_id WHERE a.paper_id=? AND a.status='submitted' ORDER BY a.score_earned DESC""", (pid,)).fetchall()]
        for r in rows: r["ratio"] = r["score_earned"]/r["max_score"] if r["max_score"] else 0
    return jsonify({"paper": dict(p), "q_stats": q_stats, "score_distribution": ranges, "submissions": rows})

def _get_local_ips():
    """获取本机所有 IPv4 地址"""
    import socket
    ips = []
    try:
        hostname = socket.gethostname()
        for info in socket.getaddrinfo(hostname, None, family=socket.AF_INET):
            ip = info[4][0]
            if not ip.startswith("127.") and ip not in ips:
                ips.append(ip)
    except Exception: pass
    if not ips:
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("223.5.5.5", 80))
            ip = s.getsockname()[0]
            s.close()
            if ip and not ip.startswith("127."):
                ips.append(ip)
        except Exception: pass
    return ips

def _print_startup_banner():
    """启动时打印访问信息（云端通过日志显示）"""
    init_db()
    print("=" * 64)
    print("  中考英语真题题库 v2.0.5 - 服务已启动")
    print("=" * 64)
    if USE_POSTGRES:
        print("  数据库: PostgreSQL (云端)")
    else:
        print(f"  数据库: SQLite (本地) - {DB_PATH}")
    print(f"  监听端口: {PORT}")
    print("  本机访问:    http://localhost:%d" % PORT)
    for ip in _get_local_ips():
        print(f"  局域网访问:  http://{ip}:{PORT}   <- 学生在浏览器打开这个")
    print("  默认账号:    admin / admin123  (首次登录后请改密码)")
    print("  停止服务:    按 Ctrl+C")
    print("=" * 64)

# 让 gunicorn 可以直接 import 'app:app'（无需 if __name__ == "__main__"）
# 云端通过 gunicorn 启动时不会执行 if __name__ 块
# 所以 init_db 放到第一次请求时（懒加载），或者 gunicorn 启动前手动 init

if __name__ == "__main__":
    _print_startup_banner()
    app.run(host="0.0.0.0", port=PORT, debug=False, use_reloader=False, threaded=True)